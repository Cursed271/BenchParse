# ----- License -------------------------------------------------- # 

#  BenchParse - BenchParse converts CIS Benchmark PDF reports into structured Excel files for efficient compliance analysis and tracking.
#  Copyright (c) 2025 - CursedSec (Operated by Cursed271). All rights reserved.

#  This software is an proprietary intellectual property developed for
#  penetration testing, threat modeling, and security research. It   
#  is licensed under the CURSEDSEC OWNERSHIP EDICT:
#
#  🚫 PROHIBITION WARNING 🚫
#  Redistribution, re-uploading, and unauthorized modification are strictly forbidden 
#  under the COE. Use is granted ONLY under the limited terms defined in the official 
#  LICENSE file (COE), which must be included in all copies.

#  DISCLAIMER:
#  This tool is intended for **educational or ethical testing** purposes only.
#  Unauthorized or malicious use of this software against systems without 
#  proper authorization is strictly prohibited and may violate laws and regulations.
#  The author assumes no liability for misuse or damage caused by this tool.

#  🔗 LICENSE: CURSEDSEC OWNERSHIP EDICT (COE)
#  🔗 Repository: https://github.com/Cursed271
#  🔗 Author: Steven Pereira (@Cursed271)

# ----- Libraries ------------------------------------------------------ #

import re
import os
import pdfplumber
from halo import Halo
from rich.console import Console
from rich.text import Text
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl import Workbook

# ----- Global Declaration --------------------------------------------- #

console = Console()
spinner = Halo(color="green")
#recommendation_pattern = re.compile(r'^\s*(\d+(?:\.\d+)+)\s+(.+)')
remove_pattern = re.compile(r'Page\s\d{1,3}|•')
title_pattern = re.compile(r'^(\d+\.\d+(?:\.\d+)*)\s*(\(L\d+\))?\s*(.*)')
page_number_pattern = re.compile(r'\bPage\s+\d+\b', re.IGNORECASE)
sections = {
    'Description': 'Description:',
    'Rationale': 'Rationale:',
    'Impact': 'Impact:',
    'Audit': 'Audit:',
    'Profile Applicability': 'Profile Applicability:',
    'Remediation': 'Remediation:'
}

# ----- Banner --------------------------------------------------------- #

def banner():
    console.print(rf"""[#C6ECE3]
┌───────────────────────────────────────────────────────────────────────────────────────────────────────────────────┐
│                                                                                                                   │
│     oooooooooo.                                  oooo        ooooooooo.                                           │ 
│    `888'   `Y8b                                 `888        `888   `Y88.                                          │
│     888     888  .ooooo.  ooo. .oo.    .ooooo.   888 .oo.    888   .d88'  .oooo.   oooo d8b  .oooo.o  .ooooo.     │
│     888oooo888' d88' `88b `888P"Y88b  d88' `"Y8  888P"Y88b   888ooo88P'  `P  )88b  `888""8P d88(  "8 d88' `88b    │
│     888    `88b 888ooo888  888   888  888        888   888   888          .oP"888   888     `"Y88b.  888ooo888    │
│     888    .88P 888    .o  888   888  888   .o8  888   888   888         d8(  888   888     o.  )88b 888    .o    │
│    o888bood8P'  `Y8bod8P' o888o o888o `Y8bod8P' o888o o888o o888o        `Y888""8o d888b    8""888P' `Y8bod8P'    │
│                                                                                                                   │
└───────────────────────────────────────────────────────────────────────────────────────────────────────────────────┘
	""")
    console.print("[#C6ECE3]+--------------------------------------------------------------+")
    console.print("[#C6ECE3]  BenchParse - Parse. Analyze. Export. CIS Made Simple.")
    console.print("[#C6ECE3]  Created by [bold black]Cursed271")
    console.print("[#C6ECE3]+--------------------------------------------------------------+")

# ----- Remove Page Numbers -------------------------------------------- #

def remove_pgno(text):
    return page_number_pattern.sub('', text)

# ----- Save to Excel -------------------------------------------------- #

def save_output(recommendations, output_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "CIS Benchmark"
    header_font = Font(name = "Aptos", size = 12, bold = True, color = "000000")
    header_fill = PatternFill(start_color = "00B0F0", end_color = "00B0F0", fill_type = "solid")
    cell_font = Font(name = "Aptos", size = 12, color = "000000")
    cell_fill = PatternFill(start_color = "FFFFFF", end_color = "FFFFFF", fill_type = "solid")
    headers = ['Control Name', 'Control Title', 'Description', 'Rationale', 'Impact', 'Audit', 'Recommendation', 'Profile Applicability']
    sheet.append(headers)
    for col_num in range(1, len(headers) + 1):
        cell = sheet.cell(row = 1, column = col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal = 'left')
    for recommendation in recommendations:
        row_values = [
            recommendation.get('Number', ''),
            recommendation.get('Title', ''),
            recommendation.get('Description', ''),
            recommendation.get('Rationale', ''),
            recommendation.get('Impact', ''),
            recommendation.get('Audit', ''),
            recommendation.get('Remediation', ''),
            recommendation.get('Profile Applicability', '')
        ]
        sheet.append(row_values)
    for row_num in range(2, len(recommendations) + 2):
        for col_num in range(1, len(headers) + 1):
            cell = sheet.cell(row = row_num, column = col_num)
            cell.font = cell_font
            cell.fill = cell_fill
            cell.alignment = Alignment(horizontal = 'left')
    sheet.column_dimensions['A'].width = 15
    sheet.column_dimensions['B'].width = 35
    sheet.column_dimensions['C'].width = 35
    sheet.column_dimensions['D'].width = 35
    sheet.column_dimensions['E'].width = 35
    sheet.column_dimensions['F'].width = 35
    sheet.column_dimensions['G'].width = 35
    sheet.column_dimensions['H'].width = 35
    num_rows = len(recommendations)
    num_cols = len(headers)
    tab_range = f"A1:{chr(64 + num_cols)}{num_rows + 1}"
    tab = Table(displayName = "Benchmark", ref = tab_range)
    style = TableStyleInfo(
        name = "TableStyleMedium2", 
        showFirstColumn = False, 
        showLastColumn = False, 
        showRowStripes = False, 
        showColumnStripes = False, 
    )
    tab.tableStyleInfo = style
    sheet.add_table(tab)
    workbook.save(output_path)

# ----- Read CIS Benchmark --------------------------------------------- #

def read_pdf(input_path):
    text = []
    with pdfplumber.open(input_path) as pdf:
        total_pages = len(pdf.pages)
        extraction_started = False
        for page_number, page in enumerate(pdf.pages[9:], start = 10):
            page_text = page.extract_text()
            if not extraction_started:
                if "Recommendations" in page_text and "....." not in page_text and "Recommendation Definitions" not in page_text:
                    extraction_started = True
            if extraction_started:
                if "Appendix: Summary Table" in page_text or "Checklist" in page_text:
                    break
                text.append(page_text)
    return '\n'.join(text)

# ----- Extract Profile Applicability ---------------------------------- #

def extract_profile(lines, start_index, max_depth = 10):
    for i in range(start_index + 1, min(start_index + max_depth, len(lines))):
        line = lines[i].strip()
        if line.startswith("Profile Applicability:"):
            return True
        if title_pattern.match(line) or any(line.startswith(sec) for sec in sections.values()):
            return False
    return False

# ----- Extract Recommendation ----------------------------------------- #

def extract_recommend(text):
    recommendations = []
    lines = text.splitlines()
    current_recommendation = {}
    current_index = 0
    while current_index < len(lines):
        line = lines[current_index].strip()
        line = remove_pgno(line)
        title_match = title_pattern.match(line)
        if title_match:
            if extract_profile(lines, current_index):
                if current_recommendation:
                    recommendations.append(current_recommendation)
                current_recommendation = {
                    'Number': title_match.group(1),
                    'Title': title_match.group(3),
                }
                while (
                    current_index + 1 < len(lines) and 
                    not any (lines[current_index + 1].strip().startswith(sec) for sec in sections.values()) and
                    not title_pattern.match(lines[current_index + 1].strip())
                ):
                    current_index += 1
                    current_recommendation['Title'] += " " + lines[current_index].strip()
        for section_name, section_start in sections.items():
            if line.startswith(section_start):
                content, next_index = extract_section(lines, current_index)
                current_recommendation[section_name] = content
                current_index = next_index - 1
                break
        current_index += 1
    if current_recommendation:
        recommendations.append(current_recommendation)
    unique_recommendations = {(rec['Number'], rec['Title']): rec for rec in recommendations}
    return list(unique_recommendations.values())

# ----- Extract Section ------------------------------------------------ #

def extract_section(lines, start_index):
    content = []
    current_index = start_index + 1
    while current_index < len(lines):
        line = lines[current_index].strip()
        line = remove_pgno(line)
        if any(line.startswith(sec) for sec in sections.values()) or title_pattern.match(line) or 'CIS Controls' in line:
            break
        if line.lower().startswith("references:") or line.lower().startswith("default value:"):
            break
        content.append(line)
        current_index += 1
    cleaned_content = ' '.join(content).strip()
    return cleaned_content, current_index

# ----- Menu ----------------------------------------------------------- #

def menu():
    input_path = console.input("[#C6ECE3][?] Enter the Benchmark that will be converted to Excel: ")
    output_path = console.input("[#C6ECE3][?] Enter the name of the Output File: ")
    spinner.start()
    text = read_pdf(input_path)
    recommendations = extract_recommend(text)
    save_output(recommendations, output_path)
    console.print(f"[green][+] Finished completed the Benchmark to Excel - {output_path}")
    console.print("[#C6ECE3]+--------------------------------------------------------------+")

# ----- Main Function -------------------------------------------------- #

if __name__ == "__main__":
    os.system("cls" if os.name == "nt" else "clear")
    banner()
    menu()

# ----- End ------------------------------------------------------------ #
